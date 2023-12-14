import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, timedelta

class operator(object):
    def __init__(self, GUI):
        self.revord_path = GUI.revord_path.get()
        self.ship_to_path = GUI.ship_to_path.get()
        self.sold_to_path = GUI.sold_to_path.get()
        self.allocation_path = GUI.allocation_path.get()
        self.dn_path = GUI.dn_path.get()
        self.stock_path = GUI.stock_path.get()
        self.customer_priority_path = GUI.customer_priority_path.get()
        self.exception_path = GUI.exception_path.get()
        self.save_path = GUI.save_path.get()
        
        # memory the last working day and holiday
        self.last_working_day = GUI.calendar.selected_last_working_day
        self.holiday = GUI.calendar.selected_holiday
        
        # this excel file is '.xlsx' format, so we can use pandas to read it directly
        self.revord_df = pd.read_excel(self.revord_path)
        self.dn_df = pd.read_excel(self.dn_path)
        
        # this excel file is '.xls' format, so we need to read it by pandas
        self.sold_to_df = pd.read_excel(self.sold_to_path)
        self.ship_to_df = pd.read_excel(self.ship_to_path)
        self.allocation_df = pd.read_excel(self.allocation_path)
        self.CPN_df = pd.read_excel(self.allocation_path, sheet_name=1)
        self.stock_df = pd.read_excel(self.stock_path)
        self.customer_priority_df = pd.read_excel(self.customer_priority_path)
        self.exception_df = pd.read_excel(self.exception_path)
    
    def sold_to_check(self):
        # Convert 'Sold To No.' column in revord_df to string and remove leading zeros
        self.revord_df['Sold To No.'] = self.revord_df['Sold To No.'].astype(str).str.lstrip('0')

        # Add a new column 'shipping point', 'CPN', 'DDL block' with default value '' and 'EETT', 'ETT' with 0
        self.revord_df['shipping point'] = ''
        self.revord_df['CPN'] = ''
        self.revord_df['EETT'] = 0
        self.revord_df['ETT'] = 0
        self.revord_df['DDL block'] = None
        self.revord_df['Stock'] = 0
        self.revord_df['Proposed PGI'] = ''
        self.revord_df['Remark'] = ''
        self.revord_df['leaf seller'] = ''
        self.revord_df['Arrange stock'] = ''
        
        # add a new column 'Plant2' and fill it with the value of 'Plant' removing the number
        self.revord_df['Plant2'] = self.revord_df['Plant'].str.replace(r'\d+', '', regex=True)

        # Convert 'SoldTo' column in sold_to_df to string for consistent comparison
        self.sold_to_df['SoldTo'] = self.sold_to_df['SoldTo'].astype(str)

        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Check if modified 'Sold To No.' is in 'SoldTo' column of sold_to_df
            if row['Sold To No.'] in self.sold_to_df['SoldTo'].values:
                # Update 'DDL block' column
                self.revord_df.at[index, 'DDL block'] = 'sold to block'
    
    def ship_to_check(self):
        # Convert 'Ship To No.' column in revord_df to string and remove leading zeros
        self.revord_df['Ship To No.'] = self.revord_df['Ship To No.'].astype(str).str.lstrip('0')

        # Convert 'ShipTo' column in ship_to_df to string for consistent comparison
        self.ship_to_df['ShipTo'] = self.ship_to_df['ShipTo'].astype(str)

        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Check if modified 'Ship To No.' is in 'ShipTo' column of ship_to_df
            if row['Ship To No.'] in self.ship_to_df['ShipTo'].values:
                # Check if 'DDL block' column already has a value
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    # Concatenate new comment with existing comment
                    self.revord_df.at[index, 'DDL block'] += '; ship to block'
                else:
                    # Update 'DDL block' column with new comment
                    self.revord_df.at[index, 'DDL block'] = 'ship to block'
    
    def allocation_check(self):
        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Prepare the comment to be added
            comment_to_add = f"{row['Material entered']} {row['Plant']} block"

            # Check if there is a matching row in allocation_df
            if any((self.allocation_df['SP'] == row['Material entered']) & (self.allocation_df['Plant'] == row['Plant'])):
                # Check if 'DDL block' column already has a value
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    # Concatenate new comment with existing comment
                    self.revord_df.at[index, 'DDL block'] += '; ' + comment_to_add
                else:
                    # Update 'DDL block' column with new comment
                    self.revord_df.at[index, 'DDL block'] = comment_to_add
    
    def add_dn_infro(self):
        # Iterate through each row in revord_df
        for index, revord_row in self.revord_df.iterrows():
            # Find matching rows in dn_df
            matching_rows = self.dn_df[
                (self.dn_df['Sales Doc.'] == revord_row['Sales Document']) &
                (self.dn_df['Item'] == revord_row['Sales Document Item'])
            ]

            # If there's a match, update the relevant columns in revord_df
            if not matching_rows.empty:
                # Assuming the first matching row is the relevant one
                matching_row = matching_rows.iloc[0]

                # Update 'shipping point', 'EETT', 'ETT', 'CPN' in revord_df with values from dn_df
                self.revord_df.at[index, 'shipping point'] = matching_row['ShPt']
                # check nan, padding with 0
                if pd.notna(matching_row['EETT']):
                    self.revord_df.at[index, 'EETT'] = matching_row['EETT']
                if pd.notna(matching_row['ETT']):
                    self.revord_df.at[index, 'ETT'] = matching_row['ETT']
                self.revord_df.at[index, 'CPN'] = matching_row['Customer Material Number']
                # update leaf seller and modify the value of 'Seller H' column. It is like 'ABB_4051490_WF00::LEVEL3'. We need the string before '::'
                if pd.notna(matching_row['Seller H']):
                    self.revord_df.at[index, 'leaf seller'] = matching_row['Seller H'].split('::')[0]
                else:
                    self.revord_df.at[index, 'leaf seller'] = ''
        # convert 'EETT' and 'ETT' to int type, they are like 1,00. We need the int 1
        for index, row in self.revord_df.iterrows():
            if type(row['EETT']) == str:
                self.revord_df.at[index, 'EETT'] = int(row['EETT'].split(',')[0])
            if type(row['ETT']) == str:
                self.revord_df.at[index, 'ETT'] = int(row['ETT'].split(',')[0])
        # add a new column 'transit' wihch is the result of 'EETT' + 'ETT'
        self.revord_df['Transit'] = self.revord_df['EETT'] + self.revord_df['ETT']
    
    def dn_check(self):
        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Prepare the comment to be added
            comment_to_add = 'CPN+Plant block'

            # Get the first and second column names of CPN_df
            first_column = self.CPN_df.columns[0]
            # the second column removed the number
            self.CPN_df[self.CPN_df.columns[1]] = self.CPN_df[self.CPN_df.columns[1]].str.replace(r'\d+', '', regex=True)

            second_column = self.CPN_df.columns[1]

            # Check if there is a matching row in self.CPN_df
            if any((self.CPN_df[first_column] == row['CPN']) & (self.CPN_df[second_column] == row['Plant2'])):
                # Check if 'DDL block' column already has a value
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    # Concatenate new comment with existing comment
                    self.revord_df.at[index, 'DDL block'] += '; ' + comment_to_add
                else:
                    # Update 'DDL block' column with new comment
                    self.revord_df.at[index, 'DDL block'] = comment_to_add
    
    def exception_check(self):
        comment_to_add = 'exception handling'
        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Check if there is a matching row in exception_df with the column 'sold-to', 'ship-to', 'MC', 'CPN', 'SP'
            # the corresponding coulumn in revord_df is 'Sold To No.', 'Ship To No.', 'Main Customer', 'CPN', 'Material entered'
            # if the row has one value in the column, it will add the comment to the 'DDL block' column
            # only check the not nan value in exception_df
            if pd.notna(row['Sold To No.']) and any(self.exception_df['sold-to'] == row['Sold To No.']):
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    self.revord_df.at[index, 'DDL block'] += '; ' + comment_to_add
                else:
                    self.revord_df.at[index, 'DDL block'] = comment_to_add
            if pd.notna(row['Ship To No.']) and any(self.exception_df['ship-to'] == row['Ship To No.']):
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    self.revord_df.at[index, 'DDL block'] += '; ' + comment_to_add
                elif comment_to_add in self.revord_df.at[index, 'DDL block']:
                    continue
                else:
                    self.revord_df.at[index, 'DDL block'] = comment_to_add
            if pd.notna(row['Main Customer']) and any(self.exception_df['MC'] == row['Main Customer']):
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    self.revord_df.at[index, 'DDL block'] += '; ' + comment_to_add
                elif comment_to_add in self.revord_df.at[index, 'DDL block']:
                    continue
                else:
                    self.revord_df.at[index, 'DDL block'] = comment_to_add
            if pd.notna(row['CPN']) and any(self.exception_df['CPN'] == row['CPN']):
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    self.revord_df.at[index, 'DDL block'] += '; ' + comment_to_add
                elif comment_to_add in self.revord_df.at[index, 'DDL block']:
                    continue
                else:
                    self.revord_df.at[index, 'DDL block'] = comment_to_add
            if pd.notna(row['Material entered']) and any(self.exception_df['SP'] == row['Material entered']):
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    self.revord_df.at[index, 'DDL block'] += '; ' + comment_to_add
                elif comment_to_add in self.revord_df.at[index, 'DDL block']:
                    continue
                else:
                    self.revord_df.at[index, 'DDL block'] = comment_to_add

    def add_stock(self):
        # Ensure that 'Stock' column exists in revord_df
        if 'Stock' not in self.revord_df.columns:
            self.revord_df['Stock'] = 0

        # Iterate through each row in revord_df
        for index, revord_row in self.revord_df.iterrows():
            # Find matching rows in stock_df
            matching_rows = self.stock_df[
                (self.stock_df['Plnt'] == revord_row['Plant']) &
                (self.stock_df['Material'] == revord_row['Material'])
            ]

            # If there's a match, update the 'Stock' value in revord_df
            if not matching_rows.empty:
                # Assuming the first matching row is the relevant one
                matching_row = matching_rows.iloc[0]

                # Update 'Stock' in revord_df with 'Sum of stock' from stock_df
                self.revord_df.at[index, 'Stock'] = matching_row['FREEQTY']
                
    def cal_date(self, date, dw):
        # Convert the string date to a datetime object for comparison
        if isinstance(date, pd.Timestamp):
            start_date = date.date()
        elif isinstance(date, str):
            start_date = datetime.strptime(date, '%Y-%m-%d').date()
        else:
            start_date = date
        # Assuming self.holiday is a list of holidays in 'YYYY-MM-DD' format
        holidays = [datetime.strptime(day, '%Y-%m-%d').date() for day in self.holiday]
        days_to_subtract = dw
        while days_to_subtract > 0:
            start_date -= timedelta(days=1)
            # Skip weekends and holidays
            if start_date.weekday() >= 5 or start_date in holidays:
                continue
            else:
                days_to_subtract -= 1
        return start_date
    
    def cal_proposed_day(self):
        # Convert the string date to a datetime object for comparison
        last_working_day = datetime.strptime(self.last_working_day[0], '%Y-%m-%d').date()
        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Convert 'Goods Issue Date' to datetime.date for comparison
            if pd.isnull(row['Goods Issue Date']):
                continue  # or handle the missing value as needed
            elif isinstance(row['Goods Issue Date'], pd.Timestamp):
                goods_issue_date = row['Goods Issue Date'].date()
            else:
                # Convert 'Goods Issue Date' to datetime.date for comparison if it's a string
                goods_issue_date = datetime.strptime(row['Goods Issue Date'], '%Y-%m-%d').date()
            # convert 'Customer requested date' to datetime.date for comparison
            if pd.isnull(row['Customer requested date']):
                continue
            elif isinstance(row['Customer requested date'], pd.Timestamp):
                crd = row['Customer requested date'].date()
            else:
                crd = datetime.strptime(row['Customer requested date'], '%Y-%m-%d').date()
            
            # the EETT and ETT is like '1,00' and '2,00', convert it to int 1, 2...
            if pd.isnull(row['Del Window Minus']):
                dw = 0
            else:
                dw = int(row['Del Window Minus'])
            if type(row['EETT']) == str:
                eett = int(row['EETT'].split(',')[0])
            elif type(row['EETT']) == int:
                pass
            else:
                eett = 0
            if type(row['ETT']) == str:
                ett = int(row['ETT'].split(',')[0])
            elif type(row['ETT']) == int:
                pass
            else:
                ett = 0

            # Check the conditions and update 'Remark' and 'Proposed PGI' accordingly
            if goods_issue_date <= last_working_day:
                self.revord_df.at[index, 'Remark'] = 'Open AT'
                self.revord_df.at[index, 'Proposed PGI'] = goods_issue_date
            else:
                if self.cal_date(goods_issue_date, dw) <= last_working_day:
                    if self.cal_date(date=crd, dw=ett+eett) <= last_working_day:
                        self.revord_df.at[index, 'Remark'] = 'Open AT'
                        self.revord_df.at[index, 'Proposed PGI'] = self.cal_date(goods_issue_date, dw) # last_working_day - dw
                    else:
                        self.revord_df.at[index, 'Remark'] = 'DW potential'
                        self.revord_df.at[index, 'Proposed PGI'] = self.cal_date(goods_issue_date, dw) # goods_issue_date - dw
                else:
                    if self.cal_date(date=crd, dw=ett+eett) <= last_working_day:
                        self.revord_df.at[index, 'Remark'] = 'Due CRD with late GI'
                        self.revord_df.at[index, 'Proposed PGI'] = self.cal_date(date=crd, dw=ett+eett) # crd - ett - eett
                    else:
                        if self.cal_date(date=crd, dw=ett+eett) <= last_working_day:
                            self.revord_df.at[index, 'Remark'] = 'CRD potential with late GI'
                            self.revord_df.at[index, 'Proposed PGI'] = self.cal_date(date=crd, dw=ett+eett) # crd - ett - eett
                        else:
                            self.revord_df.at[index, 'Remark'] = 'No potential'
                            self.revord_df.at[index, 'Proposed PGI'] = None
    
    def arrange_stock(self):
        # 添加默认值为0的新列
        self.revord_df['Priority'] = 1
        self.revord_df['sum of value'] = 0

        # 设置优先级并计算'sum of value'
        for index, row in self.revord_df.iterrows():
            matching_rows = self.customer_priority_df[
                (self.customer_priority_df['Sales Product #'] == row['Material entered']) & 
                (self.customer_priority_df['Leaf Seller'] == row['leaf seller'])
            ]
            if not matching_rows.empty:
                priority = matching_rows.iloc[0]['Calculated JIRA Prio']
                self.revord_df.at[index, 'Priority'] = priority
            self.revord_df.at[index, 'sum of value'] = row['Priority'] * row['Net Value In EUR']

        # 定义'Remark'的自定义排序函数
        def remark_sorter(remark):
            priorities = {
                'Open AT': 1,
                'DW potential': 2,
                'Due CRD with late GI': 3,
                'CRD potential with late GI': 4
            }
            return priorities.get(remark, 5)  # 其他备注的默认优先级

        # 按'Material entered'和'Plant'对数据进行分组
        grouped = self.revord_df.groupby(['Material entered', 'Plant'])

        for name, group in grouped:
            # 首先按'Remark'优先级排序，然后按'sum of value'排序
            sorted_group = group.sort_values(by=['Remark', 'sum of value'], key=lambda x: x.map(remark_sorter) if x.name == 'Remark' else x, ascending=[True, False])

            # 初始化rest_stock
            rest_stock = sorted_group.iloc[0]['Stock']

            for index, row in sorted_group.iterrows():
                if pd.notna(row['DDL block']) or row['Stock'] == 0 or row['Remark'] == "No potential":
                    self.revord_df.at[index, 'Arrange stock'] = -1
                else:
                    if rest_stock > row['Open Quantity']:
                        # 可以出货
                        self.revord_df.at[index, 'Arrange stock'] = 1
                        rest_stock -= row['Open Quantity']
                    else:
                        # 不能出货
                        self.revord_df.at[index, 'Arrange stock'] = 0

    def save(self):
        # Format dates in the DataFrame
        self.revord_df['Customer requested date'] = self.revord_df['Customer requested date'].dt.strftime('%Y/%m/%d')
        self.revord_df['Goods Issue Date'] = self.revord_df['Goods Issue Date'].dt.strftime('%Y/%m/%d')
        self.revord_df['Delivery Date'] = self.revord_df['Delivery Date'].dt.strftime('%Y/%m/%d')

        # Define the Excel file path
        file_path = os.path.join(self.save_path, 'original data.xlsx')

        # Use a context manager to handle the ExcelWriter
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write the DataFrame to an Excel file
            self.revord_df.to_excel(writer, sheet_name='RevOrd', index=False)

            # Access the workbook and sheet for formatting
            worksheet = writer.sheets['RevOrd']

            # Set the column width
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

            # Set header style
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
        
        # find the 'arrange stock' == 1
        self.df = self.revord_df[self.revord_df['Arrange stock'] == 1]
        # only save the columns we need
        columns = ['Sales Office', 'Description', 'Sold To No.', 'Ship To No.', 'Sales Document Type', 'CPN', 'leaf seller', 'Material entered', 'Material', 'shipping point', 'Sales Document', 'Sales Document Item', 'Open Quantity', 'Customer requested date', 'Transit', 'Allocation policy', 'Open Quantity', 'Proposed PGI']
        df_template = self.df[columns]
        # write the template to excel
        file_path_template = os.path.join(self.save_path, 'infineon revenue.xlsx')
                # Use a context manager to handle the ExcelWriter
        with pd.ExcelWriter(file_path_template, engine='openpyxl') as writer:
            # Write the DataFrame to an Excel file
            df_template.to_excel(writer, sheet_name='RevOrd', index=False)

            # Access the workbook and sheet for formatting
            worksheet = writer.sheets['RevOrd']

            # Set the column width with every column width
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

            # Set header style
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
        