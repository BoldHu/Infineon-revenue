import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, timedelta
class operator(object):
    def __init__(self):
        # self.revord_path = 'Rev activity/ZSD_REVORD 20231115.XLSX'
        self.revord_path = 'Rev activity/add_stock.xlsx'
        self.ship_to_path = 'Rev activity/DDue_List_ShipTo_Data 20231115.xls'
        self.sold_to_path = 'Rev activity/DDue_List_SoldTo_Data 20231115.xls'
        self.allocation_path = 'Rev activity/DDue_List_Allocation_Data 20231115.xls'
        self.dn_path = 'Rev activity/ZSD_CHIPBIZ_DN 2023115.XLSX'
        self.stock_path = 'Rev activity/Stock via SQ01.XLSX'
        self.save_path = 'Rev activity/'
        
        # this excel file is '.xlsx' format, so we can use pandas to read it directly
        self.revord_df = pd.read_excel(self.revord_path)
        self.dn_df = pd.read_excel(self.dn_path)
        
        # this excel file is '.xls' format, so we need to read it by pandas
        self.sold_to_df = pd.read_excel(self.sold_to_path)
        self.ship_to_df = pd.read_excel(self.ship_to_path)
        self.allocation_df = pd.read_excel(self.allocation_path)
        self.CPN_df = pd.read_excel(self.allocation_path, sheet_name=1)
        self.stock_df = pd.read_excel(self.stock_path)
        
        self.last_working_day = ['2023-12-29']
        self.holiday = ['2023-12-25']
    
    def sold_to_check(self):
        # Convert 'Sold To No.' column in revord_df to string and remove leading zeros
        self.revord_df['Sold To No.'] = self.revord_df['Sold To No.'].astype(str).str.lstrip('0')

        # Add a new column 'shipping point', 'CPN', 'DDL block' with default value '' and 'EETT', 'ETT' with 0
        self.revord_df['shipping point'] = ''
        self.revord_df['CPN'] = ''
        self.revord_df['EETT'] = 0
        self.revord_df['ETT'] = 0
        self.revord_df['DDL block'] = ''
        self.revord_df['Stock'] = 0
        self.revord_df['Proposed PGI'] = ''
        self.revord_df['Remark'] = ''
        self.revord_df['Arrange stock'] = ''

        # Convert 'SoldTo' column in sold_to_df to string for consistent comparison
        self.sold_to_df['SoldTo'] = self.sold_to_df['SoldTo'].astype(str)

        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Check if modified 'Sold To No.' is in 'SoldTo' column of sold_to_df
            if row['Sold To No.'] in self.sold_to_df['SoldTo'].values:
                # Update 'DDL block' column
                self.revord_df.at[index, 'DDL block'] = 'sold to block'

        # save the result to excel
        self.revord_df.to_excel('Rev activity/sold_to_check.xlsx', index=False)
        print("Sold-to check completed!")
    
    def ship_to_check(self):
        # Convert 'Ship To No.' column in revord_df to string and remove leading zeros
        self.revord_df['Ship To No.'] = self.revord_df['Ship To No.'].astype(str).str.lstrip('0')

        # Convert 'ShipTo' column in ship_to_df to string for consistent comparison
        self.ship_to_df['ShipTo'] = self.ship_to_df['ShipTo'].astype(str)

        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Check if modified 'Ship To No.' is in 'ShipTo' column of ship_to_df
            if row['Ship To No.'] in self.ship_to_df['ShipTo'].values:
                # Check if 'DDL block' column already has a value
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    # Concatenate new comment with existing comment
                    self.revord_df.at[index, 'DDL block'] += '; ship to block'
                else:
                    # Update 'DDL block' column with new comment
                    self.revord_df.at[index, 'DDL block'] = 'ship to block'
        # save the result to excel
        self.revord_df.to_excel('Rev activity/ship_to_check.xlsx', index=False)
        print("Ship-to check completed!")
    
    def allocation_check(self):
        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Prepare the comment to be added
            comment_to_add = f"{row['Material entered']} {row['Plant']} block"

            # Check if there is a matching row in allocation_df
            if any((self.allocation_df['SP'] == row['Material entered']) & (self.allocation_df['Plant'] == row['Plant'])):
                # Check if 'DDL block' column already has a value
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    # Concatenate new comment with existing comment
                    self.revord_df.at[index, 'DDL block'] += '; ' + comment_to_add
                else:
                    # Update 'DDL block' column with new comment
                    self.revord_df.at[index, 'DDL block'] = comment_to_add
                    
        # save the result to excel
        self.revord_df.to_excel('Rev activity/allocation_check.xlsx', index=False)
        print("Allocation check completed!")
        
    def add_dn_infro(self):
        # Iterate through each row in revord_df
        for index, revord_row in self.revord_df.iterrows():
            # Find matching rows in dn_df
            matching_rows = self.dn_df[
                (self.dn_df['Sales Doc.'] == revord_row['Sales Document']) &
                (self.dn_df['Item'] == revord_row['Sales Document Item'])
            ]

            # If there's a match, update the relevant columns in revord_df
            if not matching_rows.empty:
                # Assuming the first matching row is the relevant one
                matching_row = matching_rows.iloc[0]

                # Update 'shipping point', 'EETT', 'ETT', 'CPN' in revord_df with values from dn_df
                self.revord_df.at[index, 'shipping point'] = matching_row['ShPt']
                # check nan, padding with 0
                if pd.notna(matching_row['EETT']):
                    self.revord_df.at[index, 'EETT'] = matching_row['EETT']
                if pd.notna(matching_row['ETT']):
                    self.revord_df.at[index, 'ETT'] = matching_row['ETT']
                self.revord_df.at[index, 'CPN'] = matching_row['Customer Material Number']
        # save result
        self.revord_df.to_excel('Rev activity/add_dn_infro.xlsx', index=False)
        print('DN information added!')
    
    def dn_check(self):
        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Prepare the comment to be added
            comment_to_add = 'CPN+Plant block'

            # Get the first and second column names of CPN_df
            first_column = self.CPN_df.columns[0]
            second_column = self.CPN_df.columns[1]

            # Check if there is a matching row in self.CPN_df
            if any((self.CPN_df[first_column] == row['CPN']) & (self.CPN_df[second_column] == row['Plant'])):
                # Check if 'DDL block' column already has a value
                if pd.notna(self.revord_df.at[index, 'DDL block']) and self.revord_df.at[index, 'DDL block'] != '':
                    # Concatenate new comment with existing comment
                    self.revord_df.at[index, 'DDL block'] += '; ' + comment_to_add
                else:
                    # Update 'DDL block' column with new comment
                    self.revord_df.at[index, 'DDL block'] = comment_to_add
        # save the result
        self.revord_df.to_excel('Rev activity/dn_check.xlsx')
        print('dn check')
    
    def add_stock(self):
        # Ensure that 'Stock' column exists in revord_df
        if 'Stock' not in self.revord_df.columns:
            self.revord_df['Stock'] = 0

        # Iterate through each row in revord_df
        for index, revord_row in self.revord_df.iterrows():
            # Find matching rows in stock_df
            matching_rows = self.stock_df[
                (self.stock_df['SP'] == revord_row['Material entered']) &
                (self.stock_df['Plnt'] == revord_row['Plant']) &
                (self.stock_df['Material'] == revord_row['Material'])
            ]

            # If there's a match, update the 'Stock' value in revord_df
            if not matching_rows.empty:
                # Assuming the first matching row is the relevant one
                matching_row = matching_rows.iloc[0]

                # Update 'Stock' in revord_df with 'Sum of stock' from stock_df
                self.revord_df.at[index, 'Stock'] = matching_row['FREEQTY']
        # save the result
        self.revord_df.to_excel('Rev activity/add_stock.xlsx', index=False)
        print('add stock completed!')
    
    def cal_proposed_day(self):
        def cal_date(self, date, dw):
            # Convert the string date to a datetime object for comparison
            start_date = datetime.strptime(date, '%Y-%m-%d').date()
            # Assuming self.holiday is a list of holidays in 'YYYY-MM-DD' format
            holidays = [datetime.strptime(day, '%Y-%m-%d').date() for day in self.holiday]
            days_to_subtract = dw
            while days_to_subtract > 0:
                start_date -= timedelta(days=1)
                # Skip weekends and holidays
                if start_date.weekday() >= 5 or start_date in holidays:
                    continue
                else:
                    days_to_subtract -= 1
            return start_date
        
        # Convert the string date to a datetime object for comparison
        last_working_day = datetime.strptime(self.last_working_day[0], '%Y-%m-%d').date()
        # Iterate through each row in revord_df
        for index, row in self.revord_df.iterrows():
            # Convert 'Goods Issue Date' to datetime.date for comparison
            if pd.isnull(row['Goods Issue Date']):
                continue  # or handle the missing value as needed
            elif isinstance(row['Goods Issue Date'], pd.Timestamp):
                goods_issue_date = row['Goods Issue Date'].date()
            else:
                # Convert 'Goods Issue Date' to datetime.date for comparison if it's a string
                goods_issue_date = datetime.strptime(row['Goods Issue Date'], '%Y-%m-%d').date()
            # convert 'Customer requested date' to datetime.date for comparison
            if pd.isnull(row['Customer requested date']):
                continue
            elif isinstance(row['Customer requested date'], pd.Timestamp):
                crd = row['Customer requested date'].date()
            else:
                crd = datetime.strptime(row['Customer requested date'], '%Y-%m-%d').date()
            
            # the EETT and ETT is like '1,00' and '2,00', convert it to int 1, 2...
            if pd.isnull(row['Del Window Minus']):
                dw = 0
            else:
                dw = int(row['Del Window Minus'])
            if type(row['EETT']) == str:
                eett = int(row['EETT'].split(',')[0])
            if type(row['ETT']) == str:
                ett = int(row['ETT'].split(',')[0])

            # Check the conditions and update 'Remark' and 'Proposed PGI' accordingly
            if goods_issue_date <= last_working_day:
                self.revord_df.at[index, 'Remark'] = 'Open AT'
                self.revord_df.at[index, 'Proposed PGI'] = goods_issue_date
            else:
                if cal_date(goods_issue_date, dw) <= last_working_day:
                    if crd-ett-eett <= last_working_day:
                        self.revord_df.at[index, 'Remark'] = 'Open AT'
                        self.revord_df.at[index, 'Proposed PGI'] = cal_date(last_working_day, dw) # last_working_day - dw
                    else:
                        self.revord_df.at[index, 'Remark'] = 'DW potential'
                        self.revord_df.at[index, 'Proposed PGI'] = cal_date(goods_issue_date, dw) # goods_issue_date - dw
                else:
                    if crd - ett - eett <= last_working_day:
                        self.revord_df.at[index, 'Remark'] = 'Due CRD with late GI'
                        self.revord_df.at[index, 'Proposed PGI'] = cal_date(date=crd, dw=ett+eett) # crd - ett - eett
                    else:
                        if cal_date(date=crd, dw=ett+eett) <= last_working_day:
                            self.revord_df.at[index, 'Remark'] = 'CRD potential with late GI'
                            self.revord_df.at[index, 'Proposed PGI'] = cal_date(date=crd, dw=ett+eett) # crd - ett - eett
                        else:
                            self.revord_df.at[index, 'Remark'] = 'No potential'
                            self.revord_df.at[index, 'Proposed PGI'] = None
        # save the result
        self.revord_df.to_excel('Rev activity/cal_proposed_day.xlsx', index=False)
        print('Proposed PGI added!')
    
    def arrange_stock(self):
        pass

    def save(self):
        # Format dates in the DataFrame
        self.revord_df['Customer requested date'] = self.revord_df['Customer requested date'].dt.strftime('%Y/%m/%d')
        self.revord_df['Goods Issue Date'] = self.revord_df['Goods Issue Date'].dt.strftime('%Y/%m/%d')
        self.revord_df['Delivery Date'] = self.revord_df['Delivery Date'].dt.strftime('%Y/%m/%d')

        # Define the Excel file path
        file_path = os.path.join(self.save_path, 'infineon revenue.xlsx')

        # Use a context manager to handle the ExcelWriter
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write the DataFrame to an Excel file
            self.revord_df.to_excel(writer, sheet_name='RevOrd', index=False)

            # Access the workbook and sheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['RevOrd']

            # Set the column width
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

            # Set header style
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
    
if  __name__ == '__main__':
    operator = operator()
    # operator.sold_to_check()
    # print("Sold-to check completed!")
    # operator.ship_to_check()
    # print("Ship-to check completed!")
    # operator.allocation_check()
    # print("Allocation check completed!")
    # operator.add_dn_infro()
    # print("DN information added!")
    # operator.dn_check()
    # print("DN check completed!")
    # operator.add_stock()
    # print("Stock added!")
    operator.cal_proposed_day()
    print("Proposed PGI added!")
    operator.arrange_stock()
    print("Stock arranged!")
    operator.save()
    print("Report saved!")
    print("Report generated successfully!")
        